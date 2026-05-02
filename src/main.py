import time
import sqlite3
from datetime import datetime

import cv2
import numpy as np
import flet as ft
import flet_camera as fc
from openpyxl import load_workbook

DB = "attendance.db"
ADMIN_PASSWORD = "123456"
EXCEL_PATH_KEY = "qr_attendance.excel_path"
QR_SEPARATOR = "|"

qr_detector = cv2.QRCodeDetector()


# -------------------- DATABASE --------------------
def init_db():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS students (
            reg_no TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            name_norm TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reg_no TEXT NOT NULL,
            name TEXT NOT NULL,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            UNIQUE(reg_no, date)
        )
    """)

    conn.commit()
    conn.close()


def normalize_text(text: str) -> str:
    return " ".join(str(text).strip().lower().split())


def replace_students(students: list[tuple[str, str]]):
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("DELETE FROM students")
    cur.executemany(
        "INSERT INTO students (reg_no, name, name_norm) VALUES (?, ?, ?)",
        [(reg_no, name, normalize_text(name)) for reg_no, name in students],
    )
    conn.commit()
    conn.close()


def get_students():
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("SELECT reg_no, name FROM students ORDER BY reg_no")
    rows = cur.fetchall()
    conn.close()
    return rows


def count_students() -> int:
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM students")
    result = cur.fetchone()
    conn.close()
    return int(result[0]) if result else 0


def lookup_student(reg_no: str):
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("SELECT name, name_norm FROM students WHERE reg_no = ?", (reg_no,))
    row = cur.fetchone()
    conn.close()
    return row


def already_marked_today(reg_no: str) -> bool:
    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute(
        "SELECT 1 FROM attendance WHERE reg_no = ? AND date = ? LIMIT 1",
        (reg_no, today),
    )
    exists = cur.fetchone() is not None
    conn.close()
    return exists


def mark_attendance(reg_no: str, name: str):
    now = datetime.now()
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute(
        "INSERT OR IGNORE INTO attendance (reg_no, name, date, time) VALUES (?, ?, ?, ?)",
        (reg_no, name, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")),
    )
    conn.commit()
    conn.close()


def get_today_attendance():
    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT time, reg_no, name
        FROM attendance
        WHERE date = ?
        ORDER BY time DESC
        """,
        (today,),
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def count_today_attendance() -> int:
    today = datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM attendance WHERE date = ?", (today,))
    result = cur.fetchone()
    conn.close()
    return int(result[0]) if result else 0


# -------------------- EXCEL IMPORT --------------------
def clean_header(v):
    if v is None:
        return ""
    return str(v).strip().lower().replace(" ", "").replace("_", "")


def import_students_from_excel(path: str) -> list[tuple[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    students: list[tuple[str, str]] = []

    # Detect header row if present
    header_row = None
    header_map = {}

    for i, row in enumerate(rows[:5]):
        normalized = [clean_header(cell) for cell in row]

        has_name = any(x in {"name", "fullname", "studentname", "الاسم", "اسم"} for x in normalized)
        has_reg = any(x in {"regno", "reg_no", "reg", "registrationno", "رقمالقيد", "القيد", "رقم"} for x in normalized)

        if has_name and has_reg:
            header_row = i
            for idx, cell in enumerate(normalized):
                if cell in {"name", "fullname", "studentname", "الاسم", "اسم"}:
                    header_map["name"] = idx
                if cell in {"regno", "reg_no", "reg", "registrationno", "رقمالقيد", "القيد", "رقم"}:
                    header_map["reg_no"] = idx
            break

    if header_row is not None:
        name_idx = header_map.get("name", 0)
        reg_idx = header_map.get("reg_no", 1)

        for row in rows[header_row + 1:]:
            if not row:
                continue

            name = ""
            reg_no = ""

            if name_idx < len(row) and row[name_idx] is not None:
                name = str(row[name_idx]).strip()

            if reg_idx < len(row) and row[reg_idx] is not None:
                reg_no = str(row[reg_idx]).strip()

            if name and reg_no:
                students.append((reg_no, name))
    else:
        for row in rows:
            if not row or len(row) < 2:
                continue

            name = str(row[0]).strip() if row[0] is not None else ""
            reg_no = str(row[1]).strip() if row[1] is not None else ""

            if name and reg_no:
                students.append((reg_no, name))

    unique = {}
    for reg_no, name in students:
        unique[reg_no] = name

    return [(reg_no, name) for reg_no, name in unique.items()]


# -------------------- APP --------------------
async def main(page: ft.Page):
    page.title = "نظام التحضير بالـ QR"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.theme = ft.Theme(use_material3=True, color_scheme_seed=ft.Colors.INDIGO)
    page.padding = 0
    page.scroll = ft.ScrollMode.AUTO
    page.horizontal_alignment = ft.CrossAxisAlignment.STRETCH

    prefs = page.shared_preferences

    state = {
        "is_admin": False,
        "camera_ready": False,
        "last_qr": "",
        "last_qr_ts": 0.0,
    }

    def notify(text: str):
        page.snack_bar = ft.SnackBar(content=ft.Text(text))
        page.snack_bar.open = True
        page.update()

    # ---------- Controls ----------
    total_students_text = ft.Text("0", size=28, weight=ft.FontWeight.BOLD)
    total_today_text = ft.Text("0", size=28, weight=ft.FontWeight.BOLD)
    scan_result = ft.Text("وجّه الكاميرا نحو الـ QR", size=16, weight=ft.FontWeight.BOLD)

    students_list = ft.Column(spacing=10)
    today_list = ft.Column(spacing=10)

    password_input = ft.TextField(
        label="كلمة مرور الأدمن",
        password=True,
        can_reveal_password=True,
        border_radius=14,
        width=320,
    )

    excel_path_input = ft.TextField(
        label="مسار ملف Excel",
        hint_text="/storage/emulated/0/Download/students.xlsx",
        border_radius=14,
        width=420,
    )

    excel_status = ft.Text("", size=13, color=ft.Colors.GREY_700)

    camera_preview = fc.Camera(
        expand=True,
        preview_enabled=True,
        content=ft.Container(
            alignment=ft.alignment.center,
            content=ft.Icon(ft.Icons.CENTER_FOCUS_STRONG, color=ft.Colors.WHITE70, size=54),
        ),
    )

    # ---------- Helpers ----------
    def refresh_stats():
        total_students_text.value = str(count_students())
        total_today_text.value = str(count_today_attendance())
        page.update()

    def refresh_students():
        students_list.controls.clear()
        rows = get_students()

        if not rows:
            students_list.controls.append(
                ft.Container(
                    padding=16,
                    border_radius=16,
                    bgcolor=ft.Colors.GREY_100,
                    content=ft.Text("لا يوجد طلاب بعد. استورد ملف Excel."),
                )
            )
        else:
            for reg_no, name in rows:
                students_list.controls.append(
                    ft.Container(
                        padding=16,
                        border_radius=16,
                        bgcolor=ft.Colors.WHITE,
                        border=ft.border.all(1, ft.Colors.BLUE_GREY_100),
                        content=ft.Row(
                            [
                                ft.Column(
                                    [
                                        ft.Text(name, size=16, weight=ft.FontWeight.BOLD),
                                        ft.Text(reg_no, size=12, color=ft.Colors.GREY_600),
                                    ],
                                    spacing=2,
                                ),
                                ft.Icon(ft.Icons.PERSON, color=ft.Colors.INDIGO),
                            ],
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        ),
                    )
                )

        refresh_stats()

    def refresh_today():
        today_list.controls.clear()
        rows = get_today_attendance()

        if not rows:
            today_list.controls.append(
                ft.Container(
                    padding=16,
                    border_radius=16,
                    bgcolor=ft.Colors.GREY_100,
                    content=ft.Text("لا يوجد حضور مسجل اليوم."),
                )
            )
        else:
            for time_str, reg_no, name in rows:
                today_list.controls.append(
                    ft.Container(
                        padding=16,
                        border_radius=16,
                        bgcolor=ft.Colors.WHITE,
                        border=ft.border.all(1, ft.Colors.BLUE_GREY_100),
                        content=ft.Row(
                            [
                                ft.Row(
                                    [
                                        ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN),
                                        ft.Column(
                                            [
                                                ft.Text(name, weight=ft.FontWeight.BOLD),
                                                ft.Text(reg_no, size=12, color=ft.Colors.GREY_600),
                                            ],
                                            spacing=1,
                                        ),
                                    ],
                                    spacing=8,
                                ),
                                ft.Text(time_str, color=ft.Colors.GREY_700),
                            ],
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        ),
                    )
                )

        refresh_stats()

    def parse_qr_text(qr_text: str):
        qr_text = qr_text.strip()
        if QR_SEPARATOR not in qr_text:
            return None, None

        name, reg_no = qr_text.split(QR_SEPARATOR, 1)
        name = name.strip()
        reg_no = reg_no.strip()

        if not name or not reg_no:
            return None, None

        return name, reg_no

    def process_scanned_qr(qr_text: str):
        scanned_name, reg_no = parse_qr_text(qr_text)

        if not reg_no:
            scan_result.value = "صيغة QR غير صحيحة. المطلوب: الاسم|رقم القيد"
            page.update()
            return

        record = lookup_student(reg_no)
        if not record:
            scan_result.value = f"الطالب غير موجود: {reg_no}"
            page.update()
            return

        db_name, db_name_norm = record

        if normalize_text(scanned_name) != db_name_norm:
            scan_result.value = f"اسم QR لا يطابق Excel: {db_name}"
            page.update()
            return

        if already_marked_today(reg_no):
            scan_result.value = f"{db_name} مسجل مسبقًا اليوم"
            page.update()
            return

        mark_attendance(reg_no, db_name)
        scan_result.value = f"تم تسجيل حضور: {db_name}"
        refresh_today()
        page.update()

    async def load_saved_excel_path():
        saved_path = await prefs.get(EXCEL_PATH_KEY)
        if saved_path:
            excel_path_input.value = str(saved_path)
            page.update()

    async def save_excel_path(e=None):
        path = excel_path_input.value.strip()
        if not path:
            excel_status.value = "أدخل المسار أولًا."
            page.update()
            return

        await prefs.set(EXCEL_PATH_KEY, path)
        excel_status.value = "تم حفظ المسار."
        notify("تم حفظ مسار Excel")
        page.update()

    async def import_excel_from_path(e=None):
        path = excel_path_input.value.strip()
        if not path:
            excel_status.value = "أدخل مسار ملف Excel."
            page.update()
            return

        try:
            students = import_students_from_excel(path)
            if not students:
                excel_status.value = "الملف فارغ أو غير صالح."
                page.update()
                return

            replace_students(students)
            excel_status.value = f"تم استيراد {len(students)} طالبًا."
            notify("تم استيراد الطلاب من Excel")
            refresh_students()
            refresh_today()
            page.update()
        except Exception as ex:
            excel_status.value = f"خطأ في الاستيراد: {ex}"
            page.update()

    def on_stream_image(e: fc.CameraImageEvent):
        try:
            now = time.monotonic()

            # throttle repeated reads
            if now - state["last_qr_ts"] < 1.0:
                return

            frame = np.frombuffer(e.bytes, dtype=np.uint8)
            image = cv2.imdecode(frame, cv2.IMREAD_COLOR)
            if image is None:
                return

            decoded, points, _ = qr_detector.detectAndDecode(image)
            if not decoded:
                return

            if decoded == state["last_qr"] and (now - state["last_qr_ts"]) < 2.0:
                return

            state["last_qr"] = decoded
            state["last_qr_ts"] = now
            process_scanned_qr(decoded)

        except Exception as ex:
            scan_result.value = f"خطأ في القراءة: {ex}"
            page.update()

    camera_preview.on_stream_image = on_stream_image

    async def init_camera():
        try:
            cameras = await camera_preview.get_available_cameras()
            if not cameras:
                scan_result.value = "لا توجد كاميرا متاحة"
                page.update()
                return

            selected = cameras[0]
            for cam in cameras:
                try:
                    if getattr(cam, "lens_direction", None) == fc.CameraLensDirection.BACK:
                        selected = cam
                        break
                except Exception:
                    pass

            await camera_preview.initialize(selected)

            if await camera_preview.supports_image_streaming():
                await camera_preview.start_image_stream()
                state["camera_ready"] = True
                scan_result.value = "الكاميرا جاهزة للمسح"
                page.update()
            else:
                scan_result.value = "بث الصور غير مدعوم على هذا الجهاز"
                page.update()

        except Exception as ex:
            scan_result.value = f"فشل تشغيل الكاميرا: {ex}"
            page.update()

    async def stop_camera():
        if state["camera_ready"]:
            try:
                await camera_preview.stop_image_stream()
            except Exception:
                pass
            state["camera_ready"] = False

    def header():
        return ft.Container(
            padding=20,
            bgcolor=ft.Colors.INDIGO,
            content=ft.Column(
                [
                    ft.Text("نظام التحضير QR", size=26, weight=ft.FontWeight.BOLD, color=ft.Colors.WHITE),
                    ft.Text("استيراد الطلاب من Excel ثم التحضير بالكاميرا", size=14, color=ft.Colors.WHITE),
                    ft.Row(
                        [
                            ft.OutlinedButton("الحضور", on_click=lambda e: page.run_task(show_attendance)),
                            ft.OutlinedButton("الإدارة", on_click=lambda e: page.run_task(show_admin_gate)),
                        ],
                        spacing=12,
                    ),
                ],
                spacing=8,
            ),
        )

    def stat_cards():
        return ft.Row(
            [
                ft.Container(
                    expand=1,
                    padding=16,
                    border_radius=18,
                    bgcolor=ft.Colors.WHITE,
                    border=ft.border.all(1, ft.Colors.BLUE_GREY_100),
                    content=ft.Column(
                        [ft.Text("عدد الطلاب"), total_students_text],
                        spacing=6,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                ),
                ft.Container(
                    expand=1,
                    padding=16,
                    border_radius=18,
                    bgcolor=ft.Colors.WHITE,
                    border=ft.border.all(1, ft.Colors.BLUE_GREY_100),
                    content=ft.Column(
                        [ft.Text("حضور اليوم"), total_today_text],
                        spacing=6,
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                ),
            ],
            spacing=12,
        )

    def card(title, content):
        return ft.Container(
            padding=16,
            border_radius=18,
            bgcolor=ft.Colors.WHITE,
            border=ft.border.all(1, ft.Colors.BLUE_GREY_100),
            content=ft.Column([ft.Text(title, size=18, weight=ft.FontWeight.BOLD), content], spacing=12),
        )

    async def show_attendance():
        await stop_camera()
        page.controls.clear()

        page.add(
            ft.SafeArea(
                content=ft.Column(
                    [
                        header(),
                        ft.Container(
                            padding=16,
                            content=ft.Column(
                                [
                                    stat_cards(),
                                    card(
                                        "المسح المباشر",
                                        ft.Container(
                                            height=360,
                                            border_radius=24,
                                            clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
                                            bgcolor=ft.Colors.BLACK,
                                            content=camera_preview,
                                        ),
                                    ),
                                    card(
                                        "النتيجة",
                                        ft.Column(
                                            [
                                                scan_result,
                                                ft.Text(
                                                    "صيغة QR المطلوبة: الاسم|رقم القيد",
                                                    size=12,
                                                    color=ft.Colors.GREY_600,
                                                ),
                                                ft.TextButton(
                                                    "دخول الأدمن",
                                                    on_click=lambda e: page.run_task(show_admin_gate),
                                                ),
                                            ],
                                            spacing=10,
                                        ),
                                    ),
                                    card("آخر الحضور اليوم", today_list),
                                ],
                                spacing=16,
                            ),
                        ),
                    ],
                    spacing=0,
                )
            )
        )
        page.update()
        refresh_today()
        await init_camera()

    async def show_admin_login():
        await stop_camera()
        page.controls.clear()

        async def login_admin(e=None):
            if password_input.value.strip() == ADMIN_PASSWORD:
                state["is_admin"] = True
                notify("تم الدخول كأدمن")
                await show_admin()
            else:
                notify("كلمة المرور غير صحيحة")

        page.add(
            ft.SafeArea(
                content=ft.Column(
                    [
                        header(),
                        ft.Container(
                            padding=16,
                            content=card(
                                "دخول الأدمن",
                                ft.Column(
                                    [
                                        password_input,
                                        ft.ElevatedButton(
                                            "دخول",
                                            icon=ft.Icons.LOCK_OPEN,
                                            on_click=login_admin,
                                        ),
                                        ft.TextButton(
                                            "رجوع للحضور",
                                            on_click=lambda e: page.run_task(show_attendance),
                                        ),
                                    ],
                                    spacing=12,
                                ),
                            ),
                        ),
                    ],
                    spacing=0,
                )
            )
        )
        page.update()

    async def show_admin():
        await stop_camera()
        page.controls.clear()

        async def refresh_all():
            refresh_students()
            refresh_today()

        page.add(
            ft.SafeArea(
                content=ft.Column(
                    [
                        header(),
                        ft.Container(
                            padding=16,
                            content=ft.Column(
                                [
                                    card(
                                        "إعداد Excel",
                                        ft.Column(
                                            [
                                                excel_path_input,
                                                ft.Row(
                                                    [
                                                        ft.ElevatedButton(
                                                            "حفظ المسار",
                                                            icon=ft.Icons.SAVE,
                                                            on_click=save_excel_path,
                                                        ),
                                                        ft.ElevatedButton(
                                                            "تحميل الطلاب",
                                                            icon=ft.Icons.UPLOAD_FILE,
                                                            on_click=import_excel_from_path,
                                                        ),
                                                    ],
                                                    wrap=True,
                                                ),
                                                excel_status,
                                            ],
                                            spacing=10,
                                        ),
                                    ),
                                    card("قائمة الطلاب", students_list),
                                    ft.Row(
                                        [
                                            ft.OutlinedButton(
                                                "عودة للحضور",
                                                on_click=lambda e: page.run_task(show_attendance),
                                            ),
                                            ft.OutlinedButton("تحديث", on_click=lambda e: refresh_all()),
                                        ],
                                        spacing=12,
                                    ),
                                ],
                                spacing=16,
                            ),
                        ),
                    ],
                    spacing=0,
                )
            )
        )
        page.update()
        refresh_students()

    async def show_admin_gate():
        if state["is_admin"]:
            await show_admin()
        else:
            await show_admin_login()

    await load_saved_excel_path()
    await show_attendance()


init_db()
ft.app(target=main)
